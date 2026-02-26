"""
Generatore di Report per TradingAgents
Supporta PDF e Excel
"""

from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from typing import Dict, List, Any, Optional, Tuple


class ReportGenerator:
    """Generatore di report trading in PDF e Excel"""
    
    def __init__(self, output_dir: str = "./reports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        
    def generate_pdf_report(
        self, 
        ticker: str, 
        date: str, 
        decision: Dict[str, Any],
        analysis_data: Optional[Dict[str, Any]] = None,
        filename: Optional[str] = None
    ) -> str:
        """
        Genera un report PDF
        
        Args:
            ticker: Simbolo del titolo
            date: Data dell'analisi
            decision: Decisione dell'analisi
            analysis_data: Dati aggiuntivi dell'analisi
            filename: Nome del file (opzionale)
            
        Returns:
            Percorso del file generato
        """
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"report_{ticker}_{timestamp}.pdf"
            
        filepath = os.path.join(self.output_dir, filename)
        
        # Crea documento PDF
        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )
        
        # Stili
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f77b4'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=12,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        )
        
        normal_style = styles['Normal']
        normal_style.fontSize = 11
        
        # Contenuto
        elements = []
        
        # Titolo
        elements.append(Paragraph("RAPPORTO DI ANALISI TRADING", title_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Informazioni generali
        info_data = [
            ["Titolo:", ticker],
            ["Data Analisi:", date],
            ["Data Generazione Rapporto:", datetime.now().strftime("%d/%m/%Y %H:%M")],
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecf0f1')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Decisione
        elements.append(Paragraph("DECISIONE TRADING", heading_style))
        
        decision_text = str(decision).replace("'", '"')
        elements.append(Paragraph(decision_text, normal_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Dati di analisi aggiuntivi
        if analysis_data:
            elements.append(Paragraph("DATI DELL'ANALISI", heading_style))
            
            for key, value in analysis_data.items():
                if isinstance(value, dict):
                    elements.append(Paragraph(f"<b>{key}:</b>", normal_style))
                    for sub_key, sub_value in value.items():
                        elements.append(
                            Paragraph(f"  • {sub_key}: {sub_value}", normal_style)
                        )
                else:
                    elements.append(
                        Paragraph(f"<b>{key}:</b> {value}", normal_style)
                    )
                elements.append(Spacer(1, 0.1*inch))
        
        elements.append(Spacer(1, 0.3*inch))
        
        # Footer
        footer_text = "Rapporto generato da TradingAgents - Sistema di Analisi Trading Automatico"
        elements.append(Paragraph(footer_text, ParagraphStyle(
            'Footer',
            parent=normal_style,
            fontSize=9,
            textColor=colors.grey,
            alignment=TA_CENTER
        )))
        
        # Genera PDF
        doc.build(elements)
        
        return filepath
    
    def generate_excel_report(
        self,
        ticker: str,
        date: str,
        decision: Dict[str, Any],
        analysis_data: Optional[Dict[str, Any]] = None,
        filename: Optional[str] = None
    ) -> str:
        """
        Genera un report Excel
        
        Args:
            ticker: Simbolo del titolo
            date: Data dell'analisi
            decision: Decisione dell'analisi
            analysis_data: Dati aggiuntivi dell'analisi
            filename: Nome del file (opzionale)
            
        Returns:
            Percorso del file generato
        """
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"report_{ticker}_{timestamp}.xlsx"
            
        filepath = os.path.join(self.output_dir, filename)
        
        # Crea workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Analisi"
        
        # Stili
        header_fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Titolo
        ws['A1'] = "RAPPORTO DI ANALISI TRADING"
        ws['A1'].font = Font(bold=True, size=14, color="1f77b4")
        ws.merge_cells('A1:B1')
        
        ws.row_dimensions[1].height = 25
        
        # Informazioni generali
        row = 3
        ws[f'A{row}'] = "Titolo"
        ws[f'B{row}'] = ticker
        
        row += 1
        ws[f'A{row}'] = "Data Analisi"
        ws[f'B{row}'] = date
        
        row += 1
        ws[f'A{row}'] = "Data Generazione"
        ws[f'B{row}'] = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        # Formatta celle informazioni
        for i in range(3, row + 1):
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'A{i}'].fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
            ws[f'A{i}'].border = border
            ws[f'B{i}'].border = border
            
        # Decisione
        row += 2
        ws[f'A{row}'] = "DECISIONE TRADING"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="1f77b4")
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        decision_str = str(decision)
        ws[f'A{row}'] = decision_str
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 60
        
        # Dati aggiuntivi
        if analysis_data:
            row += 2
            ws[f'A{row}'] = "DATI DELL'ANALISI"
            ws[f'A{row}'].font = Font(bold=True, size=12, color="1f77b4")
            ws.merge_cells(f'A{row}:B{row}')
            
            row += 1
            for key, value in analysis_data.items():
                ws[f'A{row}'] = str(key)
                ws[f'B{row}'] = str(value)
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                row += 1
        
        # Larghezza colonne
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        
        # Salva
        wb.save(filepath)
        
        return filepath
    
    def generate_both_reports(
        self,
        ticker: str,
        date: str,
        decision: Dict[str, Any],
        analysis_data: Optional[Dict[str, Any]] = None
    ) -> Tuple[str, str]:
        """
        Genera sia report PDF che Excel
        
        Returns:
            Tuple (pdf_path, excel_path)
        """
        pdf_path = self.generate_pdf_report(ticker, date, decision, analysis_data)
        excel_path = self.generate_excel_report(ticker, date, decision, analysis_data)
        
        return pdf_path, excel_path
